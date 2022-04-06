from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, colors, Alignment
from datetime import datetime

MONTH_DICTIONARY = {
    '1': 'january',
    '2': 'february',
    '3': 'march',
    '4': 'april',
    '5': 'may',
    '6': 'jun',
    '7': 'july',
    '8': 'august',
    '9': 'september',
    '10': 'october',
    '11': 'november',
    '12': 'december',
}


class WorkWithExcel:

    def __init__(self, path, sheet_name):
        self.path = path
        self.sheet_name = sheet_name
        self.sheet = load_workbook(path)
        if not self.__is_sheet_exist(sheet_name):
            self.sheet.create_sheet(sheet_name)
            name_cell = self.sheet[sheet_name].cell(1, 1, value='Name')
            father_name_cell = self.sheet[sheet_name].cell(1, 2, value='Father Name')
            name_cell.font = Font(bold=True, size=9)
            name_cell.fill = PatternFill("solid", fgColor=colors.BLACK)
            name_cell.alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=False,
                shrink_to_fit=False
            )
            father_name_cell.font = Font(bold=True, size=9)
            father_name_cell.fill = PatternFill("solid", fgColor=colors.BLACK)
            father_name_cell.alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=False,
                shrink_to_fit=False
            )
            counter = 1
            for char in range(31):
                each_cell = self.sheet[sheet_name].cell(
                    1, counter + 2,
                    value=f'{counter} {str(sheet_name)[:3].upper()}'
                )
                each_cell.fill = PatternFill("solid", fgColor=colors.BLACK)
                each_cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=False,
                    shrink_to_fit=False
                )
                counter = counter + 1

            # row = self.sheet[sheet_name].row_dimensions[1]
            # row.font = Font(bold=True)
            # row.fill = PatternFill("solid", fgColor=colors.BLACK)
            # self.sheet[sheet_name].append({'1': 'hello'})

    def take_attendance(self, user_name, father_name):
        wb = self.sheet[self.sheet_name]
        today = datetime.now()
        if not self.__is_user_exist(user_name):
            self.__add_user_to_sheet(user_name, father_name)

        user = self.__find_user(user_name)

        h1 = wb.cell(user, today.day + 2, value='present')
        h1.font = Font(color=colors.WHITE)
        h1.fill = PatternFill("solid", fgColor='88cc00')
        self.__save_sheet()
        self.__close_sheet()

    def __find_user(self, user_name):
        wb = self.sheet[self.sheet_name]
        user_list = list(map(lambda x: x.value, wb['A']))
        return user_list.index(user_name) + 1

    def __is_user_exist(self, user_name):
        wb = self.sheet[self.sheet_name]
        user_list = list(map(lambda x: x.value, wb['A']))
        return user_list.__contains__(user_name)

    def __add_user_to_sheet(self, user_name, father_name):
        wb = self.sheet[self.sheet_name]
        wb.append({'A': user_name, 'B': father_name})

    def __is_sheet_exist(self, sheet_name):
        return self.sheet.sheetnames.__contains__(sheet_name)

    def __close_sheet(self):
        self.sheet.close()

    def __save_sheet(self):
        return self.sheet.save(self.path)
